from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADERS = ["Tarih", "Müşteri Adı", "Dosya Yolu"]


def _ensure_workbook(excel_path: Path) -> None:
    if excel_path.exists():
        return
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Kayitlar"
    ws.append(HEADERS)
    wb.save(excel_path)


def append_record(excel_path: Path, tarih: str, musteri_adi: str, dosya_yolu: str) -> None:
    _ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active

    # Header kontrolü (dosya başka yerden farklı oluşmuş olabilir)
    if ws.max_row < 1:
        ws.append(HEADERS)
    else:
        first_row = [c.value for c in ws[1]]
        if first_row[: len(HEADERS)] != HEADERS:
            ws.insert_rows(1)
            for i, header in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=i, value=header)

    ws.append([tarih, musteri_adi, dosya_yolu])
    wb.save(excel_path)

